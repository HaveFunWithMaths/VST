import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
from datetime import datetime, timedelta
import calendar
import os

def parse_bank_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=int(val))
    val_str = str(val).strip()
    try:
        return datetime(1899, 12, 30) + timedelta(days=int(float(val_str)))
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            pass
    return None

def parse_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    if s == '' or s == '-' or s == 'None':
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def copy_cell_style(src_cell, dst_cell):
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format

def process_bank_statement(statement_path, template_path, output_path):
    print("--- Reading Bank Statement ---")
    wb_statement = openpyxl.load_workbook(statement_path, data_only=True)
    ws_statement = wb_statement.active
    
    # Read transactions
    transactions = []
    rows_statement = list(ws_statement.iter_rows(values_only=True))
    headers = [str(h).strip() for h in rows_statement[0]]
    
    # Determine column indexes
    # Value Date, Description, Chq / Ref No., Amount, Dr / Cr, Balance
    idx_date = headers.index("Value Date")
    idx_desc = headers.index("Description")
    idx_ref = headers.index("Chq / Ref No.")
    idx_amt = headers.index("Amount")
    idx_drcr = headers.index("Dr / Cr")
    idx_bal = headers.index("Balance")
    
    for r_idx, r in enumerate(rows_statement[1:], start=2):
        if r[idx_date] is None:
            continue
        dt = parse_bank_date(r[idx_date])
        if dt is None:
            print(f"Warning: could not parse date in statement row {r_idx}: {r[idx_date]}")
            continue
        
        desc = str(r[idx_desc]).strip() if r[idx_desc] is not None else ""
        ref_no = str(r[idx_ref]).strip() if r[idx_ref] is not None else ""
        amount = parse_float(r[idx_amt])
        dr_cr = str(r[idx_drcr]).strip().upper()
        balance = parse_float(r[idx_bal])
        
        transactions.append({
            'date': dt,
            'description': desc,
            'ref_no': ref_no,
            'amount': amount,
            'dr_cr': dr_cr,
            'balance': balance,
            'row_index': r_idx
        })
        
    # Sort chronologically, preserving original bank statement sequence for same-day items
    transactions.sort(key=lambda x: (x['date'], x['row_index']))
    print(f"Total parsed transactions: {len(transactions)}")
    
    # Calculate starting balance dynamically
    if transactions:
        first_tx = transactions[0]
        if first_tx['dr_cr'] == 'DR':
            initial_balance = first_tx['balance'] + first_tx['amount']
        else:
            initial_balance = first_tx['balance'] - first_tx['amount']
    else:
        initial_balance = 0.0
    print(f"Calculated April 1st, 2025 starting balance: {initial_balance:,.2f}")
    
    # Open template
    print("\n--- Loading Output Template ---")
    wb_out = openpyxl.load_workbook(template_path)
    
    # Cache styles first
    sheet_styles = {}
    for s_name in wb_out.sheetnames:
        ws = wb_out[s_name]
        # Find tally row
        original_tally_row = None
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Tally":
                original_tally_row = r
                break
        
        if not original_tally_row:
            print(f"Warning: Tally row not found in {s_name}, assuming row 34/35")
            original_tally_row = 34 if s_name in ['Apr 2025', 'Jun 2025', 'Sep 2025', 'Nov 2025'] else 35
            if s_name == 'Feb 2026':
                original_tally_row = 32
        
        # Cache standard data cell styles from row 4
        data_styles = {}
        for c in range(1, 8):
            cell = ws.cell(row=4, column=c)
            data_styles[c] = {
                'font': cell.font,
                'fill': cell.fill,
                'border': cell.border,
                'alignment': cell.alignment,
                'number_format': cell.number_format
            }
            
        # Cache tally cell styles
        tally_styles = {}
        for c in range(1, 8):
            cell = ws.cell(row=original_tally_row, column=c)
            tally_styles[c] = {
                'font': cell.font,
                'fill': cell.fill,
                'border': cell.border,
                'alignment': cell.alignment,
                'number_format': cell.number_format
            }
            
        sheet_styles[s_name] = {
            'data_styles': data_styles,
            'tally_styles': tally_styles,
            'original_tally_row': original_tally_row,
            'data_row_height': ws.row_dimensions[4].height,
            'tally_row_height': ws.row_dimensions[original_tally_row].height
        }
    
    tally_rows_map = {}
    
    # Now process each month/sheet
    for s_idx, s_name in enumerate(wb_out.sheetnames):
        print(f"\nProcessing Sheet: {s_name}")
        ws = wb_out[s_name]
        styles = sheet_styles[s_name]
        
        # Parse month/year from sheet name
        parts = s_name.strip().split()
        mon_str, year_str = parts[0], parts[1]
        months_map = {
            'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9,
            'Oct': 10, 'Nov': 11, 'Dec': 12, 'Jan': 1, 'Feb': 2, 'Mar': 3
        }
        month = months_map[mon_str]
        year = int(year_str)
        
        # Filter transactions for this month
        month_txs = [tx for tx in transactions if tx['date'].year == year and tx['date'].month == month]
        # Group by day
        tx_by_day = {}
        for tx in month_txs:
            day = tx['date'].day
            if day not in tx_by_day:
                tx_by_day[day] = []
            tx_by_day[day].append(tx)
            
        num_days = calendar.monthrange(year, month)[1]
        
        current_row = 4
        
        for day in range(1, num_days + 1):
            day_txs = tx_by_day.get(day, [])
            
            if not day_txs:
                # No transactions for this day
                ws.row_dimensions[current_row].height = styles['data_row_height']
                
                # Column A: Date
                cell_a = ws.cell(row=current_row, column=1)
                cell_a.value = datetime(year, month, day)
                
                # Column B: Opening Balance
                cell_b = ws.cell(row=current_row, column=2)
                if s_name == 'Apr 2025' and current_row == 4:
                    cell_b.value = initial_balance
                elif current_row == 4:
                    # Link to previous month Tally G
                    prev_sheet = wb_out.sheetnames[s_idx - 1]
                    prev_tally = tally_rows_map[prev_sheet]
                    cell_b.value = f"='{prev_sheet}'!G{prev_tally}"
                else:
                    cell_b.value = f"=G{current_row - 1}"
                    
                # Column C: Inflow
                cell_c = ws.cell(row=current_row, column=3)
                cell_c.value = 0.0
                
                # Column D: Inflow Particulars
                cell_d = ws.cell(row=current_row, column=4)
                cell_d.value = "-"
                
                # Column E: Outflow
                cell_e = ws.cell(row=current_row, column=5)
                cell_e.value = 0.0
                
                # Column F: Outflow Particulars
                cell_f = ws.cell(row=current_row, column=6)
                cell_f.value = "-"
                
                # Column G: Closing Balance
                cell_g = ws.cell(row=current_row, column=7)
                cell_g.value = f"=B{current_row}+C{current_row}-E{current_row}"
                
                # Apply styles
                for c in range(1, 8):
                    # Create cell styling using cached properties
                    t_cell = ws.cell(row=current_row, column=c)
                    t_cell.font = copy(styles['data_styles'][c]['font'])
                    t_cell.fill = copy(styles['data_styles'][c]['fill'])
                    t_cell.border = copy(styles['data_styles'][c]['border'])
                    t_cell.alignment = copy(styles['data_styles'][c]['alignment'])
                    t_cell.number_format = styles['data_styles'][c]['number_format']
                
                current_row += 1
            else:
                # Multiple transactions (or 1)
                for tx in day_txs:
                    ws.row_dimensions[current_row].height = styles['data_row_height']
                    
                    # Column A: Date
                    cell_a = ws.cell(row=current_row, column=1)
                    cell_a.value = datetime(year, month, day)
                    
                    # Column B: Opening Balance
                    cell_b = ws.cell(row=current_row, column=2)
                    if s_name == 'Apr 2025' and current_row == 4:
                        cell_b.value = initial_balance
                    elif current_row == 4:
                        prev_sheet = wb_out.sheetnames[s_idx - 1]
                        prev_tally = tally_rows_map[prev_sheet]
                        cell_b.value = f"='{prev_sheet}'!G{prev_tally}"
                    else:
                        cell_b.value = f"=G{current_row - 1}"
                        
                    # Prepare particulars text
                    particulars = tx['description']
                    if tx['ref_no'] and tx['ref_no'] != '-' and tx['ref_no'] not in particulars:
                        particulars = f"{particulars} / {tx['ref_no']}"
                    
                    if tx['dr_cr'] == 'CR':
                        # Inflow
                        ws.cell(row=current_row, column=3).value = tx['amount']
                        ws.cell(row=current_row, column=4).value = particulars
                        ws.cell(row=current_row, column=5).value = 0.0
                        ws.cell(row=current_row, column=6).value = "-"
                    else:
                        # Outflow
                        ws.cell(row=current_row, column=3).value = 0.0
                        ws.cell(row=current_row, column=4).value = "-"
                        ws.cell(row=current_row, column=5).value = tx['amount']
                        ws.cell(row=current_row, column=6).value = particulars
                        
                    # Column G: Closing Balance
                    cell_g = ws.cell(row=current_row, column=7)
                    cell_g.value = f"=B{current_row}+C{current_row}-E{current_row}"
                    
                    # Apply styles
                    for c in range(1, 8):
                        t_cell = ws.cell(row=current_row, column=c)
                        t_cell.font = copy(styles['data_styles'][c]['font'])
                        t_cell.fill = copy(styles['data_styles'][c]['fill'])
                        t_cell.border = copy(styles['data_styles'][c]['border'])
                        t_cell.alignment = copy(styles['data_styles'][c]['alignment'])
                        t_cell.number_format = styles['data_styles'][c]['number_format']
                        
                    current_row += 1
        
        # Write Tally row
        tally_row = current_row
        tally_rows_map[s_name] = tally_row
        print(f"  Tally row is written at row {tally_row}")
        
        ws.row_dimensions[tally_row].height = styles['tally_row_height']
        
        # Column A: "Tally"
        cell_a = ws.cell(row=tally_row, column=1)
        cell_a.value = "Tally"
        
        # Column B: Opening Balance =B4
        cell_b = ws.cell(row=tally_row, column=2)
        cell_b.value = "=B4"
        
        # Column C: Inflow =SUBTOTAL(109, C4:C{tally_row-1})
        cell_c = ws.cell(row=tally_row, column=3)
        cell_c.value = f"=SUBTOTAL(109,C4:C{tally_row - 1})"
        
        # Column D: Inflow Particulars = None
        cell_d = ws.cell(row=tally_row, column=4)
        cell_d.value = None
        
        # Column E: Outflow =SUBTOTAL(109, E4:E{tally_row-1})
        cell_e = ws.cell(row=tally_row, column=5)
        cell_e.value = f"=SUBTOTAL(109,E4:E{tally_row - 1})"
        
        # Column F: Outflow Particulars = "-"
        cell_f = ws.cell(row=tally_row, column=6)
        cell_f.value = "-"
        
        # Column G: Closing Balance
        cell_g = ws.cell(row=tally_row, column=7)
        # Match standard template style
        cell_g.value = f"='{s_name}'!$B${tally_row}+'{s_name}'!$C${tally_row}-'{s_name}'!$E${tally_row}"
        
        # Apply styles
        for c in range(1, 8):
            t_cell = ws.cell(row=tally_row, column=c)
            t_cell.font = copy(styles['tally_styles'][c]['font'])
            t_cell.fill = copy(styles['tally_styles'][c]['fill'])
            t_cell.border = copy(styles['tally_styles'][c]['border'])
            t_cell.alignment = copy(styles['tally_styles'][c]['alignment'])
            t_cell.number_format = styles['tally_styles'][c]['number_format']
            
        # Delete any remaining rows from previous template
        original_max_row = ws.max_row
        if original_max_row > tally_row:
            print(f"  Deleting remaining {original_max_row - tally_row} rows at the bottom (from {tally_row + 1} to {original_max_row})")
            ws.delete_rows(tally_row + 1, original_max_row - tally_row)

        # Resize any tables in this worksheet to cover the new range from header (row 3) to Tally row
        for t_name in list(ws.tables.keys()):
            table = ws.tables[t_name]
            table.ref = f"A3:G{tally_row}"
            
    print("\n--- Saving Output Workbook ---")
    wb_out.save(output_path)
    print(f"Workbook successfully saved to: {output_path}")

if __name__ == "__main__":
    # Determine base directory as the directory containing this script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    statement_file = os.path.join(base_dir, "statement.xlsx")
    backup_template = os.path.join(base_dir, "OutputTemplate_Backup.xlsx")
    default_template = os.path.join(base_dir, "OutputTemplate.xlsx")
    
    if os.path.exists(backup_template):
        template_file = backup_template
    else:
        template_file = default_template
        
    output_file = os.path.join(base_dir, "OutputTemplate.xlsx")
    
    if not os.path.exists(statement_file):
        print(f"Error: Statement file not found at: {statement_file}")
        print("Please place your bank statement file named 'statement.xlsx' in the script directory.")
        print("You can use 'statementTemplate.xlsx' as a template for your bank statement.")
    elif not os.path.exists(template_file):
        print(f"Error: Template file not found at: {template_file}")
        print("Please ensure either 'OutputTemplate_Backup.xlsx' or 'OutputTemplate.xlsx' is present in the script directory.")
    else:
        process_bank_statement(
            statement_file,
            template_file,
            output_file
        )

